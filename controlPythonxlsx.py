import pandas as pd
from openpyxl import Workbook, workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from openpyxl.styles import NamedStyle ,PatternFill, Border, Side, Alignment, Protection, Font,colors
from datetime import timedelta,datetime
import datetime
import pymysql
import warnings
import calendar
#import locale
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pathlib
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
import os
import os.path
import shutil
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import locale
#locale.setlocale(locale.LC_TIME, 'es_ES')

connection = pymysql.connect(
    host = 'localhost',user='root',passwd='rtdefse',db='metas'
)
cursor = connection.cursor()
#FechaMetas=datetime.datetime.today().strftime('%Y%m%d') 
# Mes=datetime.datetime.today().strftime('%b').upper() 
# dia=datetime.datetime.today().strftime('%d')
# ano=datetime.datetime.today().strftime('%y') 
# ahora = datetime.datetime.now()
# restFecha = ahora - timedelta(days=1)
# menodia =str(restFecha.strftime("%d"))
# #print("Menodia:", menodia)
# FechaArchivo = (menodia+'-'+Mes+'-'+ano)
# print(FechaArchivo)

oldFechaMetas=datetime.datetime.today()+timedelta(days=-1)
FechaMetas = oldFechaMetas.strftime('%Y%m%d') 

print('Fechas Metas',FechaMetas)
today=datetime.datetime.today()+timedelta(days=-1)
FechaArchivo = today.strftime('%d-%b-%y').upper() 
print(FechaArchivo)

oldFechaNormal=datetime.datetime.today()+timedelta(days=-1)
FechaNormal = oldFechaMetas.strftime('%d-%b').upper() 
print('Fecha del Encabezado',FechaNormal)


def Metas():
    #FechaMetas=datetime.datetime.today().strftime('%Y%m%d') 


    results =[]
    sql='select  cantidad from metasdiarias where fecha ='+str(FechaMetas)+''

    cursor.execute(sql)
    resultado=cursor.fetchall()
    print(resultado)
    for registro in resultado:
        results = registro[0]
        
        return results
        

ResutlsMestas = Metas()
#print('resultado',ResutlsMestas)
def MetasAcumulada():
    #FechaMetas=datetime.datetime.today().strftime('%Y%m%d') 


    results =[]
    sql='SELECT sum(metaacumulada) FROM metas.metasdiarias where fecha ='+str(FechaMetas)+''

    cursor.execute(sql)
    resultado=cursor.fetchall()
    print(resultado)
    for registro in resultado:
        results = registro[0]
        
        return results
        

AcumuladaMetas = MetasAcumulada()
print('Meta-Acumulada',AcumuladaMetas)





# Función para obtener datos que cambian en tiempo real (simulación)

font = Font(name='Calibri',
                size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')
fill = PatternFill(fill_type=None,
                start_color='FFFFFFFF',
                end_color='FF000000')
border = Border(left=Side(border_style=None,
                          color='FF000000'),
                right=Side(border_style=None,
                           color='FF000000'),
                top=Side(border_style=None,
                         color='FF000000'),
                bottom=Side(border_style=None,
                            color='FF000000'),
                diagonal=Side(border_style=None,
                              color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,
                             color='FF000000'),
                vertical=Side(border_style=None,
                              color='FF000000'),
                horizontal=Side(border_style=None,
                               color='FF000000')
                )
alignment=Alignment(horizontal='general',
                     vertical='bottom',
                  text_rotation=0,
              wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
number_format = 'Personalizada'
protection = Protection(locked=True,
                        hidden=False)

#numberFormat = NamedStyle(name='comma_style', number_format='0')
centrado = Alignment(horizontal='center')
centrado2 = Alignment(horizontal='right')

borde_doble = Side(border_style="thin")
borde_cuadrado = Border(top=borde_doble,
                        right=borde_doble,
                        bottom=borde_doble,
                        left=borde_doble)

borde_doble2 = Side(border_style="thin")
borde_cuadrado2 = Border(
                        right=borde_doble2,
                        
                        left=borde_doble2)

rojo20 = Font(color='00FF0000', size=20)




def archivo(archivo):
 df=pd.read_excel(archivo,sheet_name=0, header=2)
 countt =df.groupby(['COD-TIPO-TRANS'])['COD-TIPO-TRANS'].count()
 for cod_tipo_trans, count in countt.items():
  
  dataframe = pd.DataFrame({'tipo_tx': countt.index, 'cantidad': countt.values}).set_index('tipo_tx')
 
 valores_a_sumar = [2, 15, 17, 19, 27, 36, 38, 43]

 dataframe_filtrado = dataframe[dataframe.index.isin(valores_a_sumar)]

 total_valores_a_sumar = dataframe_filtrado['cantidad'].sum()
 #print({total_valores_a_sumar})
 
 ##esto es para poder sacar los reversos 
 tx = df['COD-TIPO-TRANS']
 tipoM= df['TIPO-MENSAJE'] 
 dataframeN =  pd.DataFrame({'tipo_tx': tx.values, 'tipo-mensaje' :tipoM.values }).set_index('tipo-mensaje')
 valores_a_sumarRe = [420]
 dataframe_filtradoN = dataframeN[dataframeN.index.isin(valores_a_sumarRe)]
 dataframe_filtrado2 = pd.DataFrame(dataframe_filtradoN)
 valores_a_sumar1 = [2,17,19,27,36,43]
 dataframe_filtrado3 = dataframe_filtrado2[dataframe_filtrado2['tipo_tx'].isin(valores_a_sumar1)]
 sumaReverso = dataframe_filtrado3['tipo_tx'].count()
 
 Totalarchivo =total_valores_a_sumar - sumaReverso
 
 
 
 
 return Totalarchivo
 
total = Totalarchivo(archivo='Reportes/Processed/_Reporte_de_Conciliacion__'+str(FechaArchivo)+'.xlsx') 
print('Totalarchivo',total)


def archivo2(archivo):
 df=pd.read_excel(archivo,sheet_name=0, header=2)
 countt =df.groupby(['COD-TIPO-TRANS'])['COD-TIPO-TRANS'].count()
 for cod_tipo_trans, count in countt.items():
  
  dataframe = pd.DataFrame({'tipo_tx': countt.index, 'cantidad': countt.values}).set_index('tipo_tx')
 
 valores_a_sumar = [12,15,17,19,23,25,26,35,36,37,38,39,43,49,51,29]

 dataframe_filtrado = dataframe[dataframe.index.isin(valores_a_sumar)]

 total_valores_a_sumar = dataframe_filtrado['cantidad'].sum()
 #print({total_valores_a_sumar})
 tx = df['COD-TIPO-TRANS']
 tipoM= df['TIPO-MENSAJE'] 
 dataframeN =  pd.DataFrame({'tipo_tx': tx.values, 'tipo-mensaje' :tipoM.values }).set_index('tipo-mensaje')
 valores_a_sumarRe = [420]
 dataframe_filtrado = dataframeN[dataframeN.index.isin(valores_a_sumarRe)]
 dataframe_filtrado2 = pd.DataFrame(dataframe_filtrado)
 valores_a_sumar1 = [12,17,19,23,25,26,35,36,37,39,43,49,51]
 dataframe_filtrado3 = dataframe_filtrado2[dataframe_filtrado2['tipo_tx'].isin(valores_a_sumar1)]
 sumaReverso = dataframe_filtrado3['tipo_tx'].count()
 
 ###Quise dejar este encabezado por tema de no confundi el proceso wherever 
 tx = df['COD-TIPO-TRANS']
 tipoM= df['CODIGO-EMPRESA-PSP'] 
 dataframePSP =  pd.DataFrame({'tipo_tx': tx.values, 'tipo-mensaje' :tipoM.values }).set_index('tipo-mensaje')
 valores_a_sumarRePSP = [4004,4001,4002,4005,4003]
 dataframe_filtradoPSP = dataframePSP[dataframePSP.index.isin(valores_a_sumarRePSP)]
 dataframe_filtradoPSP = pd.DataFrame(dataframe_filtradoPSP)
 valores_a_sumar1PSP = [29]
 dataframe_filtrado3PSP = dataframe_filtradoPSP[dataframe_filtradoPSP['tipo_tx'].isin(valores_a_sumar1PSP)]
 sumaReversoPSP = dataframe_filtrado3PSP['tipo_tx'].count()

 totalQ = sumaReverso + sumaReversoPSP
 restrTxn = total_valores_a_sumar - totalQ
 #print({total_valores_a_sumar})
 return restrTxn
 
totalarchivo = archivo(archivo='Reportes/Processed/_Reporte_de_Conciliacion__'+str(FechaArchivo)+'.xlsx') 
print('totalarchivo',totalarchivo)


def ARCHIb(archivo):
    df=pd.read_excel(archivo,sheet_name=0, header=2)
    dfT = 0
    if len(df)==0:
   
     print('no existe datos')
    else:    
     
        countt =df.groupby(['COD-TIPO-TRANS'])['COD-TIPO-TRANS'].count()
        for cod_tipo_trans, count in countt.items():
        
            dataframe = pd.DataFrame({'tipo_tx': countt.index, 'cantidad': countt.values}).set_index('tipo_tx')
        
            valores_a_sumar = [19,27]

            dataframe_filtrado = dataframe[dataframe.index.isin(valores_a_sumar)]

            total_valores_a_sumar = dataframe_filtrado['cantidad'].sum()
        #print({total_valores_a_sumar})
            dfT  =total_valores_a_sumar
   
    return dfT 
 
ARCHIb = ARCHIb(archivo='Reportes/Processed/_Reporte_de_Conciliacion__'+str(FechaArchivo)+'.xlsx') 



def ARCHIBB(archivo):
    df=pd.read_excel(archivo,sheet_name=0, header=2)
    dfT = 0
    if len(df)==0:
   
     print('no existe datos')
    else:    
     
        countt =df.groupby(['COD-TIPO-TRANS'])['COD-TIPO-TRANS'].count()
        for cod_tipo_trans, count in countt.items():
        
            dataframe = pd.DataFrame({'tipo_tx': countt.index, 'cantidad': countt.values}).set_index('tipo_tx')
        
            valores_a_sumar = [19,27]

            dataframe_filtrado = dataframe[dataframe.index.isin(valores_a_sumar)]

            total_valores_a_sumar = dataframe_filtrado['cantidad'].sum()
        #print({total_valores_a_sumar})
            dfT  =total_valores_a_sumar
   
    return dfT 
 
ARCHIBB = ARCHIBB(archivo='Reportes/Processed/Reporte_de_Conciliacion__'+str(FechaArchivo)+'.xlsx') 



def ARCHIADD(archivo):
    df=pd.read_excel(archivo,sheet_name=0, header=2)
    dfT = 0
    if len(df)==0:
   
     print('no existe datos')
    else:    
     
        countt =df.groupby(['COD-TIPO-TRANS'])['COD-TIPO-TRANS'].count()
        for cod_tipo_trans, count in countt.items():
        
            dataframe = pd.DataFrame({'tipo_tx': countt.index, 'cantidad': countt.values}).set_index('tipo_tx')
        
            valores_a_sumar = [19,27]

            dataframe_filtrado = dataframe[dataframe.index.isin(valores_a_sumar)]

            total_valores_a_sumar = dataframe_filtrado['cantidad'].sum()
        #print({total_valores_a_sumar})
            dfT  =total_valores_a_sumar
   
    return dfT 
 
ARCHIADD = ARCHIADD(archivo='Reportes/Processed/0Reporte_de_Conciliacion__'+str(FechaArchivo)+'.xlsx') 


totlTransacciones =(total + total+ total+ total+ total)
resultentero = int(totlTransacciones)
print('totalLeido de os archivos ', resultentero)


sqlUpdate = 'update metasdiarias set ejecutada =%s where fecha =%s'
upd =(resultentero,FechaMetas)
#print(upd)
cursor.execute(sqlUpdate,(upd))    
connection.commit() 

def Acumulada():
    
    sql3='SELECT sum(ejecutada) FROM metas.metasdiarias where ejecutada not in(0)'
    cursor.execute(sql3)
    resultadoAc=cursor.fetchall()

    for registro in resultadoAc:
        ejecutadaacumulada = registro[0]
        #print('Ejecutada-Acumulada',ejecutadaacumulada)  
        
        return ejecutadaacumulada
    
        
totalAcumulada = Acumulada()
print('Ejecucion-Acumulada',totalAcumulada)
 



#primer monto de la meta
string_number = ResutlsMestas
# Remove the comma from the string
string_number_without_comma = string_number.replace(',', '')
# Convert the string without the comma to an integer
integer_value = int(string_number_without_comma)




result = resultentero / integer_value

rond = round(result,2)
porcientoMeAc = f"{int(result * 100)}%"
calculoEntero = f"{int(result*100)}"





#calculo para el totalTXN entre Acumulado de la Meta

cumplimiento2= totalAcumulada /  AcumuladaMetas


rond2 = round(cumplimiento2,2)
cumplimientoRoud = f"{float(cumplimiento2*100):.2f}%"
calculoEntero2 = f"{int(cumplimiento2*100)}"

print('cumplimiento acumulado',cumplimientoRoud)
# Obtener la fecha actual
#FechaMetas = datetime.datetime.today().strftime('%B')

locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
# Obtener la fecha actual
fecha_actual = datetime.datetime.now()
# Formatear la fecha en español
Mes = fecha_actual.strftime('%B').upper()





def obtener_datos_dinamicos():
    # En este ejemplo, simularemos una lista de datos que cambian en cada llamada.
    return [
        {'Bancos': '', 'Cantidad': 28 },
        {'Bancos': '', 'Cantidad': 32 },
        {'Bancos': '',  'Cantidad': 25},
        {'Bancos': '', 'Cantidad': 30},
        {'Bancos': '', 'Cantidad': 30},
      
    ]

# Nombre del archivo Excel
nombre_archivo_excel = 'COMPORTAMIENTO-TRX.xlsx'

# Crear un libro de trabajo de Excel
libro_trabajo = Workbook()

hoja_activa = libro_trabajo.active

for r in range(8,20):
    hoja_activa[f'C{r}'].number_format ='#,##0_);(#,##0)'

# for r in range(7,13,2):
#     hoja_activa[f'F{r}'].number_format= '#,##0_);(#,##0)'





    # Obtener datos dinámicos
    datos = obtener_datos_dinamicos()
    # Convertir los datos en un DataFrame de pandas
    df = pd.DataFrame(datos)
    
    
    # Limpiar la hoja de Excel antes de agregar nuevos datos
    hoja_activa.delete_rows(1, hoja_activa.max_row)
    hoja_activa.title='COMPORTAMIENTO TRX'
    hoja_activa.merge_cells('A3:B3')
    hoja_activa.merge_cells('A11:B11')
    hoja_activa.merge_cells('A14:B14')
    hoja_activa.merge_cells('A4:B4')
    #hoja_activa.cell(row=1, column=3, value=23)
    fecha =hoja_activa['A3']= ''+(FechaNormal)+''
    #estilos
    hoja_activa['A3'].alignment= centrado
    hoja_activa['A3'].border = borde_cuadrado
    hoja_activa['B3'].border = borde_cuadrado
    # hoja_activa['C3'].border = borde_cuadrado
    # hoja_activa['D3'].border = borde_cuadrado
    hoja_activa['A3'].font =Font(name = 'Arial', size = 12, color='00FFFFFF')
    #convinar y centrar 
    
    
    #hoja_activa['D3'].alignment= centrado
    #hoja_activa['C3'].alignment= centrado
    
       
    
    
    hoja_activa['A4']='TRANSACCIONES PROCESADAS'
    hoja_activa['A4'].alignment= centrado
    hoja_activa['A4'].border = borde_cuadrado
    hoja_activa['B4'].border = borde_cuadrado
    hoja_activa['A4'].font =Font(name = 'Arial', size = 13, color='00000000')
    # hoja_activa['C4'].border = borde_cuadrado
    # hoja_activa['D4'].border = borde_cuadrado
    
    hoja_activa['A5']=''
    hoja_activa['B5']=total
    hoja_activa['A5'].border = borde_cuadrado
    hoja_activa['B5'].border = borde_cuadrado
    hoja_activa['B5'].value = float(hoja_activa['B5'].value)
    hoja_activa['B5'].number_format = '#,##0_);(#,##0)'
    hoja_activa['B5'].alignment= centrado2
    
    hoja_activa['A6']=''
    hoja_activa['B6']=total
    hoja_activa['A6'].border = borde_cuadrado
    hoja_activa['B6'].border = borde_cuadrado
    hoja_activa['B6'].value = float(hoja_activa['B6'].value)
    hoja_activa['B6'].number_format = '#,##0_);(#,##0)'
    hoja_activa['B6'].alignment= centrado2
    
    hoja_activa['A7']=''
    hoja_activa['B7']=total
    hoja_activa['A7'].border = borde_cuadrado
    hoja_activa['B7'].border = borde_cuadrado
    hoja_activa['B7'].alignment= centrado2
    hoja_activa['B7'].value = float(hoja_activa['B7'].value)
    hoja_activa['B7'].number_format = '#,##0_);(#,##0)'
    
    
    hoja_activa['A8']=''
    hoja_activa['B8']=total
    hoja_activa['A8'].border = borde_cuadrado
    hoja_activa['B8'].border = borde_cuadrado
    hoja_activa['B8'].alignment= centrado2
    hoja_activa['B8'].value = float(hoja_activa['B8'].value)
    hoja_activa['B8'].number_format = '#,##0_);(#,##0)'
    
    
    hoja_activa['A9']=''
    hoja_activa['B9']=totalA
    hoja_activa['A9'].border = borde_cuadrado
    hoja_activa['B9'].border = borde_cuadrado
    hoja_activa['B9'].alignment= centrado2
    hoja_activa['B9'].value = float(hoja_activa['B9'].value)
    hoja_activa['B9'].number_format = '#,##0_);(#,##0)'
    
    
    
    hoja_activa['A10']='TOTAL'
    hoja_activa['B10']= totlTransacciones
    hoja_activa['A10'].border = borde_cuadrado
    hoja_activa['B10'].border = borde_cuadrado
    hoja_activa['B10'].font =Font(name = 'Arial', size = 10, color='00000000',bold=True)
    hoja_activa['B10'].value = float(hoja_activa['B10'].value)
    hoja_activa['B10'].number_format = '#,##0_);(#,##0)'
    
    
    hoja_activa['A11'].border = borde_cuadrado2
    hoja_activa['B11'].border = borde_cuadrado2
    hoja_activa['A14'].border = borde_cuadrado2
    hoja_activa['B14'].border = borde_cuadrado2
    
    
    
    hoja_activa['A12']='META DE DIA:'
    hoja_activa['B12']=ResutlsMestas
    hoja_activa['B12'].font =Font(name = 'Arial', size = 10, color='00000000',bold=True)
    hoja_activa['A12'].border = borde_cuadrado
    hoja_activa['B12'].border = borde_cuadrado
    hoja_activa['B12'].alignment= centrado2
    # hoja_activa['B12'].value = float(hoja_activa['B12'].value)
    # hoja_activa['B12'].number_format = '#,##0_);(#,##0)'
    
    
    hoja_activa['A13']='CUMPLIMIENTO:'
    hoja_activa['B13']=porcientoMeAc
    hoja_activa['A13'].border = borde_cuadrado
    hoja_activa['B13'].border = borde_cuadrado
    hoja_activa['B13'].alignment= centrado2
    #hoja_activa['B13'].value = float(hoja_activa['B11'].value)
    #hoja_activa['B13'].number_format = '#,##0_);(#,##0)'
    
    
    
    hoja_activa['A15']='META ACUMULADA '+(Mes)+''
    hoja_activa['B15']=AcumuladaMetas
    hoja_activa['A15'].border = borde_cuadrado
    hoja_activa['B15'].border = borde_cuadrado
    hoja_activa['B15'].alignment= centrado2
    hoja_activa['B15'].font =Font(name = 'Arial', size = 10, color='00000000',bold=True)
    hoja_activa['B15'].value = float(hoja_activa['B15'].value)
    hoja_activa['B15'].number_format = '#,##0_);(#,##0)'
    
   
   
    
    hoja_activa['A16']='ACUMULADA '+(Mes)+''
    hoja_activa['B16']=totalAcumulada
    hoja_activa['A16'].border = borde_cuadrado
    hoja_activa['B16'].border = borde_cuadrado
    hoja_activa['B16'].alignment= centrado2
    hoja_activa['B16'].font =Font(name = 'Arial', size = 10, color='00FFFFFF',bold=True)
    hoja_activa['B16'].value = float(hoja_activa['B16'].value)
    hoja_activa['B16'].number_format = '#,##0_);(#,##0)'
    
    
    hoja_activa['A17']='CUMPLMIENTO'
    hoja_activa['B17']=cumplimientoRoud
    hoja_activa['A17'].border = borde_cuadrado
    hoja_activa['B17'].border = borde_cuadrado
    hoja_activa['B17'].alignment= centrado2
    
    
    
    
    # hoja_activa['B15'].value = float(hoja_activa['B15'].value)
    # hoja_activa['B15'].number_format = '#,##0_);(#,##0)'
    
  
    
    col = hoja_activa.column_dimensions['A'].width=25
    col = hoja_activa.column_dimensions['B'].width=25
    # col.font = Font(bold=True)
    # row = hoja_activa.row_dimensions[5]
    # row.font = Font(underline="single")
       
    # col = hoja_activa.column_dimensions['B']
    # col.font = Font(bold=True)
    # row = hoja_activa.row_dimensions[5]
    # row.font = Font(underline="single")
    
    #realizar codigo para cuando el cumplimiento sea menor a 100 se rojo de lo contrario sea verde
    calculo = int(calculoEntero) 
    calculo2 = int(calculoEntero2)
    if calculo <=99:
         hoja_activa['B13'].font =Font(name = 'Arial', size = 10, color='00FF0000',bold=True)
         
    else:
        
        hoja_activa['B13'].font =Font(name = 'Arial', size = 10, color='00339966',bold=True)
    
    if calculo2 <=99:
             hoja_activa['B17'].font =Font(name = 'Arial', size = 10, color='00FF0000',bold=True)
         
    else:
        
        hoja_activa['B17'].font =Font(name = 'Arial', size = 10, color='00339966',bold=True)        
        
        
    
     #Este codigo es para rellenar las columna que desee
    colors = ['00000000']
    fillers = []
 
    for color in colors:
        temp = PatternFill(patternType='solid',
                        fgColor=color)
        fillers.append(temp)
    cell_ids = ['B16'] 
    for i in range(1):
        hoja_activa[cell_ids[i]].fill = fillers[i]    

    # Definir el estilo de borde que deseas aplicar
    TITULO = Font(
  name='Calibri',
  size=16,
  bold=True,
  italic=False,
  vertAlign=None,
  underline='none',
  strike=False,
    color='00FFFFFF')

    td =hoja_activa['B4']
    td1 =hoja_activa['B3']
    
    td.font =TITULO
    td1.font =TITULO
    celdaFecha = "000066CC"
    celdatransaccion = "00C0C0C0"  
    for rows in hoja_activa.iter_rows(min_row=3, max_row=3, min_col=1, max_col=2): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=celdaFecha, end_color=celdaFecha,fill_type = "solid")
            
    for rows in hoja_activa.iter_rows(min_row=4, max_row=4, min_col=1, max_col=2): 
        for cell in rows: 
            cell.fill = PatternFill(start_color=celdatransaccion, end_color=celdatransaccion,fill_type = "solid")        
   
    #hoja_activa['A3'].border = borde_cuadrado
    
    

    
    
 
       
         
    #hoja_activa[5:10]
    # Escribir el DataFrame en la hoja de Excel
    
    
    
    # for row in dataframe_to_rows(df, index=False, header=True):
    #     hoja_activa.append(row)
        

    # Guardar el archivo de Excel
    libro_trabajo.save(nombre_archivo_excel)

    # Esperar 5 segundos antes de actualizar los datos nuevamente (simulación)
    #time.sleep(5)
    

    # for correito in combinar1['CORREO']:
    #         for archivito in combinar1['FILE']:
    #             print("Correo: " + correito)
    #             print("Archivo: " + archivito)

def send_email():
        #print(correito)
        #print(archivito)

        email_sender = 'notificaciones@gmail'
        email_recipient =[''] 

        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['To'] =", ".join(email_recipient)
        print('email', email_recipient)
        msg['Subject'] = 'Comportamiento transaccional'
        attachment_location = nombre_archivo_excel
        email_message = 'Comportamiento transaccional'
        # f = open(each, 'rb')

        msg.attach(MIMEText(email_message, 'plain'))

        if attachment_location != '':
            filename = os.path.basename(attachment_location)
            attachment = open(attachment_location, "rb")
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition',
                            "attachment; filename= %s" % filename)
            msg.attach(part)
        # f.close()

        try:
            server = smtplib.SMTP('', 500)
            server.sendmail(email_sender, email_recipient, msg.as_string())
            print('Reporte enviado')
            server.quit()

        except smtplib.SMTPException as e:
            print(f'Error de conexion: {e}')
        return True

send_email()    


